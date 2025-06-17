from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .forms import productForm, customerForm, categoryForm, saleForm, salesDetailForm, employeeForm, supplierForm, RestockForm, UserRegistrationForm, UserLoginForm, UserUpdateForm, ExpenseForm
from .models import Product, Category, Supplier, Customer, Sale, Employee, Inventory, Sales_Details, InventoryHistory, User, Expense
from django.db import OperationalError, transaction
from django.core.exceptions import ValidationError
import logging
from django.db.models import Sum, Count
from django.utils import timezone
from datetime import datetime
import openpyxl
from django.db.models.functions import ExtractMonth

logger = logging.getLogger(__name__)

def handle_connection_error(request, error_message="Connection failed. If the problem persists, please check your internet connection or VPN"):
    logger.error(f"Connection error: {error_message}")
    return render(request, 'layout/error.html', {
        'error_message': error_message,
        'error_type': 'Connection Error'
    })

# Create your views here.
def index(request):
    return render(request,'layout/index.html')

@login_required(login_url='login')
def home(request):
    return render(request, 'layout/Home.html')

def register(request):
    if request.user.is_authenticated:
        return redirect('dashBoard')
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, '¡Registro exitoso!')
            return redirect('dashBoard')
    else:
        form = UserRegistrationForm()
    return render(request, 'layout/register.html', {'form': form})

def user_login(request):
    if request.user.is_authenticated:
        return redirect('dashBoard')
    if request.method == 'POST':
        form = UserLoginForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'¡Bienvenido {username}!')
                return redirect('dashBoard')
    else:
        form = UserLoginForm()
    return render(request, 'layout/login.html', {'form': form})

def user_logout(request):
    logout(request)
    messages.info(request, 'Has cerrado sesión exitosamente.')
    return redirect('login')

@login_required(login_url='login')
def dashBoard(request):
    try:
        products = Product.objects.all()
        sales = Sale.objects.all()
        customers = Customer.objects.all()
        categorys = Category.objects.all()
        suppliers = Supplier.objects.all()
        employees = Employee.objects.all()
        expenses = Expense.objects.all()
        cantInventory=0
        for product in products:
            cantInventory = product.amount+cantInventory

        return render(request, 'layout/dashboard.html', { 
            'productos': products, 
            'ventas': sales, 
            'clientes': customers, 
            'tipos': categorys,
            'proveedores': suppliers, 
            'empleados': employees,
            'cantInventory': cantInventory,
            'gastos' :  expenses
        })
    except OperationalError as e:
        return handle_connection_error(request)
    except Exception as e:
        logger.error(f"Unexpected error in dashboard: {str(e)}")
        return handle_connection_error(request, "An unexpected error occurred. Please try again later.")

def panel(request):
    return render(request, 'panel.html')

@login_required(login_url='login')
def sale(request):
    if request.method == 'POST':
        sale_form = saleForm(request.POST)
        if sale_form.is_valid():
            with transaction.atomic():
                venta = sale_form.save()
                productos = request.POST.getlist('producto')
                cantidades = request.POST.getlist('amount')
                precios = request.POST.getlist('unit_price')
                for prod, cant, precio in zip(productos, cantidades, precios):
                    if prod and cant and precio:
                        detalle = Sales_Details.objects.create(
                            sale=venta,
                            producto_id=prod,
                            amount=int(cant),
                            unit_price=float(precio)
                        )
                        # Descontar la cantidad vendida del producto
                        producto = Product.objects.get(id=prod)
                        producto.amount -= int(cant)
                        producto.save()

                        inventory = Inventory.objects.get(product = producto)
                        inventory.update_quantity(producto.amount)


            return redirect('sale')
    else:
        sale_form = saleForm()
        sale_form.fields['employee'].queryset = Employee.objects.filter(post__iexact='cajero')

    sales = Sale.objects.all().order_by('-date')
    return render(request, 'layout/sale.html', {
        'sale_form': sale_form,
        'sales': sales,
        'customers': Customer.objects.all(),
        'employees': Employee.objects.all(),
        'products': Product.objects.all()
    })

@login_required(login_url='login')
def product(request):
    if request.method == 'POST':
        form = productForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('product')
    else:
        form = productForm()
    products = Product.objects.all()
    return render(request, 'layout/product.html', {'form': form, 'productos': products})

    
@login_required(login_url='login')
def category(request):
    if request.method == 'POST':
        form = categoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('category')
    else:
        form = categoryForm()
    
    categories = Category.objects.all()
    return render(request, 'layout/category.html', {
        'form': form,
        'categories': categories
    })

@login_required(login_url='login')
def supplier(request):
    if request.method == 'POST':
        form = supplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('supplier')
    else:
        form = supplierForm()
    
    suppliers = Supplier.objects.all()
    return render(request, 'layout/supplier.html', {
        'form': form,
        'suppliers': suppliers
    })

@login_required(login_url='login')
def employee(request):
    if request.method == 'POST':
        form = employeeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('employee')
    else:
        form = employeeForm()
    
    employees = Employee.objects.all()
    return render(request, 'layout/employee.html', {
        'form': form,
        'employees': employees
    })

@login_required(login_url='login')
def inventory(request):
    try:
        # Obtener todos los productos que no tienen inventario
        products_without_inventory = Product.objects.filter(inventory__isnull=True)
        
        # Crear registros de inventario para productos que no lo tienen
        for product in products_without_inventory:
            Inventory.objects.create(
                product=product,
                current_quantity=product.amount
            )
        
        # Obtener todos los items de inventario
        inventory_items = Inventory.objects.all().select_related('product')
        low_stock_items = inventory_items.filter(current_quantity__lt=10)
        
        return render(request, 'layout/inventory.html', {
            'inventory_items': inventory_items,
            'low_stock_items': low_stock_items
        })
    except Exception as e:
        logger.error(f"Error en la vista de inventario: {str(e)}")
        return handle_connection_error(request, "Error al cargar el inventario. Por favor, intente nuevamente.")

@login_required(login_url='login')
def customer(request):
    if request.method == 'POST':
        form = customerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customer')
    else:
        form = customerForm()
    
    customers = Customer.objects.all()
    return render(request, 'layout/customer.html', {
        'form': form,
        'customers': customers
    })

@login_required(login_url='login')
def inventory_history(request, inventory_id):
    try:
        # Obtener el inventario y su historial
        inventory = Inventory.objects.select_related('product').get(id=inventory_id)
        history = inventory.history.all().order_by('-date')
        
        # Crear un registro inicial si no hay historial
        if not history.exists():
            InventoryHistory.objects.create(
                inventory=inventory,
                old_quantity=0,
                new_quantity=inventory.current_quantity,
                reason="Registro inicial"
            )
            history = inventory.history.all().order_by('-date')
        
        return render(request, 'layout/inventory_history.html', {
            'inventory': inventory,
            'history': history
        })
    except Inventory.DoesNotExist:
        logger.error(f"Intento de acceder a inventario inexistente: {inventory_id}")
        return handle_connection_error(request, "El producto no existe en el inventario.")
    except Exception as e:
        logger.error(f"Error al cargar el historial de inventario: {str(e)}")
        return handle_connection_error(request, "Error al cargar el historial. Por favor, intente nuevamente.")

@login_required(login_url='login')
def restock_inventory(request, inventory_id):
    try:
        inventory = Inventory.objects.select_related('product').get(id=inventory_id)
        
        if request.method == 'POST':
            form = RestockForm(request.POST)
            if form.is_valid():
                quantity = form.cleaned_data['quantity']
                reason = form.cleaned_data['reason']
                
                # Actualizar la cantidad
                new_quantity = inventory.current_quantity + quantity
                inventory.update_quantity(new_quantity, f"Reabastecimiento: {reason}")
                
                messages.success(request, f'Stock actualizado exitosamente. Nueva cantidad: {new_quantity}')
                return redirect('inventory')
        else:
            form = RestockForm()
        
        return render(request, 'layout/restock.html', {
            'form': form,
            'inventory': inventory
        })
    except Inventory.DoesNotExist:
        logger.error(f"Intento de reabastecer inventario inexistente: {inventory_id}")
        return handle_connection_error(request, "El producto no existe en el inventario.")
    except Exception as e:
        logger.error(f"Error al reabastecer inventario: {str(e)}")
        return handle_connection_error(request, "Error al reabastecer el inventario. Por favor, intente nuevamente.")
    
@login_required(login_url='login')
def editar_producto(request, producto_id):
    producto = get_object_or_404(Product, id=producto_id)
    if request.method == 'POST':
        form = productForm(request.POST, instance=producto)
        old_amount = producto.amount
        if form.is_valid():
            # Get the old amount before saving
            
            # Save the form
            form.save()
            # Get the new amount after saving
            new_amount = form.cleaned_data['amount']
            
            # Update inventory if amount changed
            if old_amount != new_amount:
                try:
                    inventory = Inventory.objects.get(product=producto)
                    inventory.update_quantity(new_amount, "Actualización desde edición de producto")
                except Inventory.DoesNotExist:
                    # Create inventory if it doesn't exist
                    Inventory.objects.create(
                        product=producto,
                        current_quantity=new_amount
                    )
            
            messages.success(request, 'Producto actualizado exitosamente.')
            return redirect('product')
    else:
        form = productForm(instance=producto)
    return render(request, 'layout/editar_producto.html', {'form': form, 'producto': producto})

@login_required(login_url='login')
def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Product, id=producto_id)
    producto.delete()
    messages.success(request, 'Producto eliminado exitosamente.')
    return redirect('product')

@login_required(login_url='login')
def editar_categoria(request, categoria_id):
    categoria = get_object_or_404(Category, id=categoria_id)
    if request.method == 'POST':
        form = categoryForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente.')
            return redirect('category')
    else:
        form = categoryForm(instance=categoria)
    return render(request, 'layout/editar_categoria.html', {'form': form, 'categoria': categoria})

@login_required(login_url='login')
def eliminar_categoria(request, categoria_id):
    categoria = get_object_or_404(Category, id=categoria_id)
    categoria.delete()
    messages.success(request, 'Categoría eliminada exitosamente.')
    return redirect('category')

@login_required(login_url='login')
def editar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Supplier, id=proveedor_id)
    if request.method == 'POST':
        form = supplierForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, 'Proveedor actualizado exitosamente.')
            return redirect('supplier')
    else:
        form = supplierForm(instance=proveedor)
    return render(request, 'layout/editar_proveedor.html', {'form': form, 'proveedor': proveedor})

@login_required(login_url='login')
def eliminar_proveedor(request, proveedor_id):
    proveedor = get_object_or_404(Supplier, id=proveedor_id)
    proveedor.delete()
    messages.success(request, 'Proveedor eliminado exitosamente.')
    return redirect('supplier')

@login_required(login_url='login')
def editar_empleado(request, empleado_id):
    empleado = get_object_or_404(Employee, id=empleado_id)
    if request.method == 'POST':
        form = employeeForm(request.POST, instance=empleado)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empleado actualizado exitosamente.')
            return redirect('employee')
    else:
        form = employeeForm(instance=empleado)
    return render(request, 'layout/editar_empleado.html', {'form': form, 'empleado': empleado})

@login_required(login_url='login')
def eliminar_empleado(request, empleado_id):
    empleado = get_object_or_404(Employee, id=empleado_id)
    empleado.delete()
    messages.success(request, 'Empleado eliminado exitosamente.')
    return redirect('employee')

@login_required(login_url='login')
def editar_cliente(request, cliente_id):
    cliente = get_object_or_404(Customer, id=cliente_id)
    if request.method == 'POST':
        form = customerForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cliente actualizado exitosamente.')
            return redirect('customer')
    else:
        form = customerForm(instance=cliente)
    return render(request, 'layout/editar_cliente.html', {'form': form, 'cliente': cliente})

@login_required(login_url='login')
def eliminar_cliente(request, cliente_id):
    cliente = get_object_or_404(Customer, id=cliente_id)
    cliente.delete()
    messages.success(request, 'Cliente eliminado exitosamente.')
    return redirect('customer')

@login_required(login_url='login')
def ver_venta(request, venta_id):
    venta = get_object_or_404(Sale, id=venta_id)
    detalles = venta.sales_details_set.all()
    return render(request, 'layout/ver_venta.html', {
        'venta': venta,
        'detalles': detalles
    })

@login_required(login_url='login')
def eliminar_venta(request, venta_id):
    venta = get_object_or_404(Sale, id=venta_id)
    if request.method == 'POST':
        # Restaurar las cantidades de productos antes de eliminar la venta
        detalles = venta.sales_details_set.all()
        for detalle in detalles:
            producto = detalle.producto
            producto.amount += detalle.amount
            producto.save()
            
            # Actualizar el inventario
            inventory = Inventory.objects.get(product=producto)
            inventory.update_quantity(producto.amount)
        
        venta.delete()
        return redirect('sale')
    detalles = venta.sales_details_set.all()
    return render(request, 'layout/confirmar_eliminar_venta.html', {
        'venta': venta,
        'detalles': detalles
    })

@login_required
def profile(request):
    return render(request, 'layout/profile.html', {
        'title': 'Mi Perfil'
    })

@login_required
def update_profile(request):
    if request.method == 'POST':
        form = UserUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu perfil ha sido actualizado exitosamente.')
            return redirect('profile')
    else:
        form = UserUpdateForm(instance=request.user)
    
    return render(request, 'layout/update_profile.html', {
        'title': 'Actualizar Perfil',
        'form': form
    })

def logout_view(request):
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente.')
    return redirect('login')

@login_required
def expenses(request):
    expenses = Expense.objects.all().order_by('-date')
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.responsible = request.user
            expense.save()
            messages.success(request, 'Gasto registrado exitosamente.')
            return redirect('expenses')
    else:
        form = ExpenseForm()
    return render(request, 'layout/expenses.html', {'form': form, 'gastos': expenses})

@login_required
def edit_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gasto actualizado exitosamente.')
            return redirect('expenses')
    else:
        form = ExpenseForm(instance=expense)
    return render(request, 'layout/edit_expense.html', {'form': form, 'expense': expense})

@login_required
def delete_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    expense.delete()
    messages.success(request, 'Gasto eliminado exitosamente.')
    return redirect('expenses')

@login_required
def reports(request):
    # Ventas
    total_ventas = Sale.objects.count()
    total_monto_ventas = Sales_Details.objects.aggregate(total=Sum('unit_price'))['total'] or 0
    ventas_por_mes = (
        Sale.objects
        .extra(select={'mes': "strftime('%%m', date)"})
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )

    # Productos
    total_productos = Product.objects.count()
    productos_bajo_stock = Product.objects.filter(amount__lt=10).count()

    # Gastos
    total_gastos = Expense.objects.count()
    total_monto_gastos = Expense.objects.aggregate(total=Sum('amount'))['total'] or 0
    gastos_por_mes = (
        Expense.objects
        .extra(select={'mes': "strftime('%%m', date)"})
        .values('mes')
        .annotate(total=Sum('amount'))
        .order_by('mes')
    )

    # Inventario
    total_inventario = Inventory.objects.aggregate(total=Sum('current_quantity'))['total'] or 0

    return render(request, 'layout/reports.html', {
        'total_ventas': total_ventas,
        'total_monto_ventas': total_monto_ventas,
        'ventas_por_mes': ventas_por_mes,
        'total_productos': total_productos,
        'productos_bajo_stock': productos_bajo_stock,
        'total_gastos': total_gastos,
        'total_monto_gastos': total_monto_gastos,
        'gastos_por_mes': gastos_por_mes,
        'total_inventario': total_inventario,
    })

@login_required
def exportar_ventas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'
    ws.append(['ID', 'Fecha', 'Cliente', 'Empleado', 'Total'])
    ventas = Sale.objects.all().order_by('-date')
    for venta in ventas:
        total = sum([d.subtotal() for d in venta.sales_details_set.all()])
        ws.append([
            venta.id,
            venta.date.strftime('%d/%m/%Y %H:%M'),
            venta.customer.name if venta.customer else '',
            venta.employee.name if venta.employee else '',
            float(total)
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_gastos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Gastos'
    ws.append(['ID', 'Fecha', 'Descripción', 'Monto', 'Responsable'])
    gastos = Expense.objects.all().order_by('-date')
    for gasto in gastos:
        ws.append([
            gasto.id,
            gasto.date.strftime('%d/%m/%Y %H:%M'),
            gasto.description,
            float(gasto.amount),
            gasto.responsible.username if gasto.responsible else ''
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=gastos.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_ventas_mes_excel(request, mes):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Ventas Mes {mes}'
    ws.append(['ID', 'Fecha', 'Cliente', 'Empleado', 'Total'])
    ventas = Sale.objects.annotate(mes=ExtractMonth('date')).filter(mes=int(mes))
    for venta in ventas:
        total = sum([d.subtotal() for d in venta.sales_details_set.all()])
        ws.append([
            venta.id,
            venta.date.strftime('%d/%m/%Y %H:%M'),
            venta.customer.name if venta.customer else '',
            venta.employee.name if venta.employee else '',
            float(total)
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=ventas_mes_{mes}.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_gastos_mes_excel(request, mes):
    import openpyxl
    from django.http import HttpResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Gastos Mes {mes}'
    ws.append(['ID', 'Fecha', 'Descripción', 'Monto', 'Responsable'])
    gastos = Expense.objects.annotate(mes=ExtractMonth('date')).filter(mes=int(mes))
    for gasto in gastos:
        ws.append([
            gasto.id,
            gasto.date.strftime('%d/%m/%Y %H:%M'),
            gasto.description,
            float(gasto.amount),
            gasto.responsible.username if gasto.responsible else ''
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=gastos_mes_{mes}.xlsx'
    wb.save(response)
    return response